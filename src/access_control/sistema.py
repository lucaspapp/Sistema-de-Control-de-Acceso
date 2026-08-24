import base64
import json
import os
import threading
import time
from datetime import datetime, timedelta
from pathlib import Path

import cv2
import paho.mqtt.client as mqtt
from openpyxl import Workbook, load_workbook
from settings import load_local_env

from face import FaceAnalyzer, obtener_snapshot, recortar_persona, guardar_historial_prohibido
from trackeo import Tracker

PROJECT_ROOT = Path(__file__).resolve().parents[2]
load_local_env()
DATA_DIR = PROJECT_ROOT / "runtime" / "access_control" / "datos"
DATA_DIR.mkdir(parents=True, exist_ok=True)
DIARIO = DATA_DIR / "estado_actual.json"
HISTORIAL = DATA_DIR / "historial.json"

EXCEL_DIR = PROJECT_ROOT / "runtime" / "access_control" / "estadisticas"
EXCEL_DIR.mkdir(parents=True, exist_ok=True)
EXCEL_FILE = EXCEL_DIR / "estadisticas.xlsx"

MQTT_HOST = os.getenv("MQTT_HOST", "127.0.0.1")
MQTT_PORT = int(os.getenv("MQTT_PORT", "7777"))
MQTT_USERNAME = os.getenv("MQTT_USERNAME", "frigate")
MQTT_PASSWORD = os.getenv("MQTT_PASSWORD", "")
MQTT_TOPIC = "frigate/events"
DETECT_REQUEST_TOPIC = "sistema/edad_genero/request"
DETECT_RESULT_TOPIC = "sistema/edad_genero/result"
DETECT_TIMEOUT = 15

face = FaceAnalyzer()
tracker = Tracker()
mqtt_client = None
lock = threading.RLock()
pending = {}


def inicio_dia(now=None):
    now = now or datetime.now()
    inicio = now.replace(hour=8, minute=0, second=0, microsecond=0)
    return inicio if now >= inicio else inicio - timedelta(days=1)


def nuevo_estado():
    inicio = inicio_dia()
    return {
        "fecha": inicio.strftime("%Y-%m-%d"),
        "inicio": inicio.isoformat(),
        "entradas": 0,
        "salidas": 0,
        "hombres": 0,
        "mujeres": 0,
        "prohibidos": 0,
        "edades": {"(15-20)": 0, "(25-32)": 0, "(38-43)": 0, "(48-53)": 0, "(60-100)": 0},
        "personas": [],
    }


estadisticas = nuevo_estado()


def guardar():
    """Guarda inmediatamente el estado actual para poder recuperarlo
    aunque el programa se cierre o se reinicie."""
    with lock:
        DIARIO.write_text(
            json.dumps(estadisticas, indent=2, ensure_ascii=False),
            encoding="utf-8"
        )


def guardar_dia_excel(estado):
    """Agrega el resumen del día cerrado al Excel."""
    try:
        if EXCEL_FILE.exists():
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Estadisticas"

            ws.append([
                "Fecha",
                "Inicio",
                "Fin",
                "Entradas",
                "Hombres",
                "Mujeres",
                "(15-20)",
                "(25-32)",
                "(38-43)",
                "(48-53)",
                "(60-100)",
                "Prohibidos"
            ])

        inicio = datetime.fromisoformat(estado["inicio"])
        fin = inicio + timedelta(days=1)

        # Evita duplicar un día si el programa se reinicia durante el cambio.
        fecha = estado["fecha"]
        ya_existe = False

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == fecha:
                ya_existe = True
                break

        if not ya_existe:
            edades = estado.get("edades", {})

            ws.append([
                fecha,
                inicio.strftime("%Y-%m-%d %H:%M:%S"),
                fin.strftime("%Y-%m-%d %H:%M:%S"),
                estado.get("entradas", 0),
                estado.get("hombres", 0),
                estado.get("mujeres", 0),
                edades.get("(15-20)", 0),
                edades.get("(25-32)", 0),
                edades.get("(38-43)", 0),
                edades.get("(48-53)", 0),
                edades.get("(60-100)", 0),
                estado.get("prohibidos", 0)
            ])

            wb.save(EXCEL_FILE)
            print(f"[EXCEL] Día {fecha} guardado en {EXCEL_FILE}")
        else:
            wb.close()
            print(f"[EXCEL] El día {fecha} ya estaba guardado.")

    except Exception as e:
        print(f"[EXCEL] Error guardando estadísticas: {e}")


def cargar_estado():
    """Recupera el estado guardado si corresponde al día estadístico actual."""
    global estadisticas

    if not DIARIO.exists():
        estadisticas = nuevo_estado()
        return

    try:
        cargado = json.loads(DIARIO.read_text(encoding="utf-8"))

        fecha_actual = inicio_dia().strftime("%Y-%m-%d")

        if cargado.get("fecha") == fecha_actual:
            estadisticas = cargado
            print(f"[DATOS] Estado recuperado: {fecha_actual}")
        else:
            # El estado pertenece a un día anterior.
            # Se guarda en Excel antes de iniciar el nuevo día.
            guardar_dia_excel(cargado)

            estadisticas = nuevo_estado()
            guardar()

            print(
                f"[DATOS] Nuevo día iniciado: "
                f"{estadisticas['fecha']}"
            )

    except Exception as e:
        print(f"[DATOS] Error recuperando estado: {e}")
        estadisticas = nuevo_estado()


def revisar_dia():
    global estadisticas

    fecha = inicio_dia().strftime("%Y-%m-%d")

    if estadisticas["fecha"] == fecha:
        return

    # Cierra el día anterior y lo guarda en Excel.
    guardar_dia_excel(estadisticas)

    # Mantiene también el historial JSON existente.
    try:
        historial = (
            json.loads(HISTORIAL.read_text(encoding="utf-8"))
            if HISTORIAL.exists()
            else []
        )
    except Exception:
        historial = []

    historial.append(estadisticas)

    HISTORIAL.write_text(
        json.dumps(historial, indent=2, ensure_ascii=False),
        encoding="utf-8"
    )

    # Comienza el nuevo día.
    estadisticas = nuevo_estado()
    guardar()

    print("===== NUEVO DÍA ESTADÍSTICO 08:00 =====")


def mostrar():
    s = estadisticas
    print("\n" + "=" * 55)
    print("                 ESTADÍSTICAS")
    print("=" * 55)
    print(f"Entradas   : {s['entradas']}")
    print(f"Salidas    : {s['salidas']}")
    print(f"Hombres    : {s['hombres']}")
    print(f"Mujeres    : {s['mujeres']}")
    for edad, cantidad in s["edades"].items():
        print(f"{edad:<10}: {cantidad}")
    print(f"Prohibidos : {s['prohibidos']}")
    print("=" * 55 + "\n")


def imagen_base64(imagen):
    ok, encoded = cv2.imencode(".jpg", imagen, [cv2.IMWRITE_JPEG_QUALITY, 85])
    if not ok:
        return None
    return base64.b64encode(encoded.tobytes()).decode("ascii")


def registrar_entrada(info, gender, age, original_age, prohibido, nombre, score):
    tid = info["tracking_id"]
    registro = {
        "tracking_id": tid,
        "timestamp": datetime.now().isoformat(),
        "camera": info["camera"],
        "tipo": "entrada",
        "genero": gender,
        "edad": age,
        "edad_modelo": original_age,
        "prohibido": prohibido,
        "persona": nombre,
        "score": score,
    }
    with lock:
        estadisticas["entradas"] += 1
        if gender == "Male":
            estadisticas["hombres"] += 1
        elif gender == "Female":
            estadisticas["mujeres"] += 1
        if age in estadisticas["edades"]:
            estadisticas["edades"][age] += 1
        if prohibido:
            estadisticas["prohibidos"] += 1
        estadisticas["personas"].append(registro)
        guardar()
    print(f"\n[ENTRADA REGISTRADA] {tid} | género={gender} | edad={age} | prohibido={prohibido}")
    mostrar()


def procesar_entrada(info):
    tid = info["tracking_id"]
    print(f"\n[ENTRADA CONFIRMADA] {tid}")

    # Primer snapshot: se usa inmediatamente para edad/género y como
    # primera muestra del reconocimiento facial.
    imagen = obtener_snapshot(info["event_id"])
    if imagen is None:
        registrar_entrada(info, None, None, None, False, None, 0.0)
        return

    persona = recortar_persona(imagen, info.get("box"))
    if persona is None:
        persona = imagen

    # --------------------------------------------------------
    # EDAD + GÉNERO: se manda al proceso Python 3.11 sin
    # esperar al reconocimiento facial.
    # --------------------------------------------------------
    encoded = imagen_base64(imagen)
    if encoded is not None:
        request = {
            "tracking_id": tid,
            "camera": info["camera"],
            "box": info.get("box"),
            "image": encoded,
            "timestamp": datetime.now().isoformat(),
        }

        with lock:
            pending[tid] = {
                "info": info,
                "prohibido": False,
                "nombre": None,
                "score": 0.0,
                "face_done": False,
                "detect_done": False,
                "gender": None,
                "age": None,
                "original_age": None,
                "created": time.time(),
            }

        mqtt_client.publish(
            DETECT_REQUEST_TOPIC,
            json.dumps(request, ensure_ascii=False),
            qos=0,
        )
        print(f"[DETECT] Solicitud enviada: {tid}")

    # --------------------------------------------------------
    # RECONOCIMIENTO FACIAL.
    # FaceAnalyzer necesita varias muestras para confirmar un
    # prohibido. Se reutiliza su misma lógica: hasta 6 muestras,
    # con el intervalo definido en face.py.
    # --------------------------------------------------------
    prohibido = False
    nombre = None
    score = 0.0

    for muestra in range(6):
        if muestra > 0:
            time.sleep(0.72)
            nueva_imagen = obtener_snapshot(info["event_id"])
            if nueva_imagen is None:
                continue
            persona = recortar_persona(nueva_imagen, info.get("box"))
            if persona is None:
                persona = nueva_imagen

        face_result = face.analizar_tracking(tid, persona)

        if face_result is None:
            continue

        nombre = face_result.get("name")
        score = face_result.get("score", 0.0)

        if face_result.get("status") == "MATCH" and face_result.get("confirmed") is True:
            prohibido = True
            break

    if prohibido:
        print(f"🚨 PERSONA PROHIBIDA: {nombre} | score={score:.3f}")
        try:
            path = guardar_historial_prohibido(persona, tid, nombre, score)
            if path:
                print(f"[FACE] Foto: {path}")
        except Exception as e:
            print(f"[FACE] Error historial: {e}")
    else:
        print(f"[FACE] No se confirmó persona prohibida para {tid}.")

    # Actualizar la información pendiente con el resultado facial.
    with lock:
        if tid in pending:
            pending[tid]["prohibido"] = prohibido
            pending[tid]["nombre"] = nombre
            pending[tid]["score"] = score
            pending[tid]["face_done"] = True

    finalizar_si_listo(tid)


def procesar_resultado_detect(data):
    tid = data.get("tracking_id")
    if not tid:
        return

    with lock:
        item = pending.get(tid)
        if item is None:
            return

        item["gender"] = data.get("gender")
        item["age"] = data.get("age")
        item["original_age"] = data.get("original_age")
        item["detect_done"] = True

    finalizar_si_listo(tid)


def finalizar_si_listo(tid):
    with lock:
        item = pending.get(tid)
        if item is None:
            return

        if not item["face_done"] or not item["detect_done"]:
            return

        pending.pop(tid, None)

    registrar_entrada(
        item["info"],
        item["gender"],
        item["age"],
        item["original_age"],
        item["prohibido"],
        item["nombre"],
        item["score"],
    )


def registrar_salida(info):
    tid = info["tracking_id"]
    with lock:
        estadisticas["salidas"] += 1
        estadisticas["personas"].append({
            "tracking_id": tid,
            "timestamp": datetime.now().isoformat(),
            "camera": info["camera"],
            "tipo": "salida",
        })
        guardar()
    print(f"\n[SALIDA] {tid}")
    mostrar()


def on_connect(client, userdata, flags, reason_code, properties=None):
    if reason_code != 0:
        print(f"[MQTT] Error: {reason_code}")
        return
    client.subscribe(MQTT_TOPIC)
    client.subscribe(DETECT_RESULT_TOPIC)
    print(f"[MQTT] Conectado {MQTT_HOST}:{MQTT_PORT}")
    print(f"[MQTT] Frigate: {MQTT_TOPIC}")
    print(f"[MQTT] Detect: {DETECT_RESULT_TOPIC}")


def on_message(client, userdata, msg):
    try:
        data = json.loads(msg.payload.decode("utf-8"))
    except Exception:
        return

    if msg.topic == DETECT_RESULT_TOPIC:
        procesar_resultado_detect(data)
        return

    if msg.topic != MQTT_TOPIC:
        return

    revisar_dia()
    info = tracker.procesar_evento(data)
    if not info:
        return

    if info["type"] == "entrada":
        threading.Thread(target=procesar_entrada, args=(info,), daemon=True).start()
    elif info["type"] == "salida":
        registrar_salida(info)


def reloj():
    while True:
        try:
            revisar_dia()
        except Exception as e:
            print(f"[RELOJ] {e}")
        time.sleep(30)


def limpiar_pending():
    while True:
        ahora = time.time()
        vencidos = []
        with lock:
            for tid, item in list(pending.items()):
                if ahora - item["created"] > DETECT_TIMEOUT:
                    vencidos.append((tid, item))
                    pending.pop(tid, None)
        for tid, item in vencidos:
            print(f"[DETECT] Timeout: {tid}. Se registra entrada con los datos disponibles.")
            registrar_entrada(
                item["info"],
                item.get("gender"),
                item.get("age"),
                item.get("original_age"),
                item["prohibido"],
                item["nombre"],
                item["score"],
            )
        time.sleep(1)


def main():
    global mqtt_client
    cargar_estado()
    threading.Thread(target=reloj, daemon=True).start()
    threading.Thread(target=limpiar_pending, daemon=True).start()

    print("=" * 60)
    print("             SISTEMA INTEGRADO")
    print("=" * 60)
    print("Python principal : entorno del sistema")
    print("Reconocimiento   : YuNet + SFace")
    print("Tracking          : Frigate")
    print("Edad/Género      : servicio Python 3.11")
    mostrar()

    mqtt_client = mqtt.Client(mqtt.CallbackAPIVersion.VERSION2)
    mqtt_client.username_pw_set(MQTT_USERNAME, MQTT_PASSWORD)
    mqtt_client.on_connect = on_connect
    mqtt_client.on_message = on_message
    print(f"Conectando a MQTT {MQTT_HOST}:{MQTT_PORT}...")
    mqtt_client.connect(MQTT_HOST, MQTT_PORT, 60)

    try:
        mqtt_client.loop_forever()
    except KeyboardInterrupt:
        print("\n[SISTEMA] Detenido.")
    finally:
        mqtt_client.disconnect()


if __name__ == "__main__":
    main()
